import os
import requests
from functools import wraps
from logging import Logger
from io import BytesIO
from typing import Optional, List, Dict, Any, Callable
from datetime import datetime
from PyPDF2 import PdfMerger
from PIL import Image
import base64
from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties
import streamlit as st
from finance_analysis.config import global_config as glob
from finance_analysis.services.logger import LoggerFactory
from finance_analysis.utils.data_models import (
    CurrencyCodeLiteral,
    CurrencyConversionOutput,
)
from finance_analysis.services.logger import LoggerFactory

my_logger = LoggerFactory(handler_type="Stream").create_module_logger()


def get_logger() -> Logger:
    """
    Retrieves a logger instance from the Streamlit session state.

    If a logger does not already exist in the session state, this function creates a new logger
    using the LoggerFactory with a stream handler and verbose output enabled, and stores it in
    the session state. Subsequent calls will return the same logger instance.

    Returns:
        Logger: The logger instance stored in the Streamlit session state.
    """
    if "logger" not in st.session_state:
        st.session_state.logger = LoggerFactory(
            handler_type="Stream", verbose=True
        ).create_module_logger()
    return st.session_state.logger


def display_logo() -> None:
    """
    Displays the company logo in the Streamlit app.

    Attempts to load and display the logo image from a predefined directory.
    If the logo file is not found, displays an error message in the Streamlit app.
    """
    try:
        logo = Image.open(f"{glob.DATA_PKG_DIR}/NemetschekGroup_White_72dpi_oRand.png")
        st.image(logo, width=350)
    except FileNotFoundError:
        st.error("Logo file not found!")


def display_pdf(file: BytesIO) -> None:
    """
    Displays a PDF file in a Streamlit app using an iframe.

    Args:
        file (BytesIO): A BytesIO object containing the PDF file data.
        width (str, optional): The width of the iframe displaying the PDF. Defaults to "100%".
        height (str, optional): The height of the iframe displaying the PDF. Defaults to "900".

    Returns:
        None: This function does not return a value. It renders the PDF in the Streamlit app.
    """
    bytes_data = file.getvalue()
    base64_pdf = base64.b64encode(bytes_data).decode("utf-8")
    pdf_display = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="700" type="application/pdf"></iframe>'
    st.markdown(pdf_display, unsafe_allow_html=True)


def display_png(file: BytesIO) -> None:
    """
    Displays a PNG file in a Streamlit app.

    Args:
        file (BytesIO): A BytesIO object containing the PNG file data.

    Returns:
        None: This function does not return a value. It renders the PNG in the Streamlit app.
    """
    try:
        # Reset the BytesIO pointer to the beginning
        file.seek(0)

        image = Image.open(file)

        # Display the image with Streamlit
        st.image(image, use_container_width=True)

    except Exception as e:
        st.error(f"Error displaying PNG: {str(e)}")


def retry(attempts: int = 3) -> Callable:
    """
    A decorator that retries the decorated function up to a specified number of times in case of an exception.

    Args:
        attempts (int): The number of attempts to retry. Defaults to 3.

    Returns:
        Callable: The decorator function that wraps the original function.

    Example:
        @retry(attempts=5)
        def example_function():
            # Function implementation
            pass

        @retry()  # Uses default 3 attempts
        def another_function():
            # Function implementation
            pass
    """

    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper_retry(*args: tuple, **kwargs: dict) -> Any:
            for attempt in range(attempts):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    my_logger.warning(f"Attempt {attempt + 1}/{attempts} failed: {e}")
                    if attempt == attempts - 1:  # Last attempt
                        my_logger.error(
                            f"❌ All {attempts} attempts failed for {func.__name__}"
                        )
                        raise

            return None  # This should never be reached

        return wrapper_retry

    return decorator


@retry(3)
def convert_currency(
    amount: float,
    from_currency: CurrencyCodeLiteral,
) -> CurrencyConversionOutput:
    """
    Convert a monetary amount from one currency to another via api.frankfurter.app

    Args:
        amount (float): Input amount to convert
        from_currency (CurrencyCodeLiteral): Abbreviation of the currency to convert from, e.g. "USD"

    Returns:
        Dict[str, float]: A dictionary containing the converted amount and the date of conversion
    """
    if from_currency == "EUR":
        return {"EUR Amount": amount, "Exchange Rate - Date": "Not Applicable"}

    # If currency is not Euro, convert to Euro:
    resp = requests.get(
        "https://api.frankfurter.app/latest",
        params={"amount": f"{amount}", "from": f"{from_currency.upper()}", "to": "EUR"},
        timeout=15,
        verify=False,
    )
    resp.raise_for_status()
    euro_amount = resp.json()["rates"].get("EUR")  # Converted amount in EUR
    date = resp.json()["date"]
    return {"EUR Amount": euro_amount, "Exchange Rate - Date": date}


def create_conversion_info(result: dict) -> str:
    """
    Create a conversion info string based on the provided result.

    Args:
        result (dict): The result dictionary containing entities.

    Returns:
        str: A formatted string with conversion information.
    """
    # Assuming the first entity contains the currency
    if not result["entities"]:
        return "No entities found."

    # Extracting the base currency from the first entity
    base_currencies = list(set(result["currencies"]))
    non_euro = [c for c in base_currencies if c != "EUR"]
    if len(non_euro) == 0:
        print("No non-Euro currencies found.")
        base_currency = "EUR"  # Default to Euro if no other currency is found
    else:
        base_currency = non_euro[0]  # Use the first non-Euro currency
        print(f"Using {base_currency} as the base currency for conversion.")

    info = convert_currency(amount=1, from_currency=base_currency)

    valid_date = datetime.now().strftime("%d.%m.%y")
    rate_info = f"Daily exchange rate: 1 {base_currency} = {info['EUR Amount']} Euro (as of {valid_date})"
    return rate_info


def merge_pdfs(
    pdf_dir: str,
    pdf_names: Optional[List[str]] = None,
    first_file: Optional[str] = None,
    output_file: str = "merged_files.pdf",
) -> None:
    """
    Merges multiple PDF files from a specified directory into a single PDF file.

    Args:
        pdf_dir (str): The directory containing the PDF files to merge.
        pdf_names (List[str]): List of PDF file names (with extension) that must be present in the directory.
        first_file (Optional[str], optional): The PDF file to be placed first in the merged output. Defaults to None.
        output_file (str, optional): The name of the output merged PDF file. Defaults to "merged_files.pdf".
    """

    all_files = os.listdir(pdf_dir)
    if pdf_names is None:
        pdf_names = [f for f in all_files if f.endswith(".pdf")]

    # Check if all pdf_names are present in directory
    missing = [f for f in pdf_names if f not in all_files]
    if missing:
        raise FileNotFoundError(
            f"The following files are missing in {pdf_dir}: {missing}"
        )

    merger = PdfMerger()

    if first_file:
        merger.merge(0, os.path.join(pdf_dir, first_file))
        print(f"Adding {first_file}...")

    for page_number, file in enumerate(pdf_names, start=1):
        if file.endswith(".pdf") and file != first_file:
            print(f"Appending {file}...")
            merger.merge(page_number, os.path.join(pdf_dir, file))

    merger.write(os.path.join(pdf_dir, output_file))
    merger.close()
    my_logger.info(f"Merged PDF created as {os.path.join(pdf_dir, output_file)}")


def update_travel_expense_xlsx(
    result: Dict[str, Any],
    trip_metadata: Dict[str, str] = {
        "Last/First name": "Vosseler, Alexander",
        "Location": "Munich",
        "Destination": "Barcelona",
        "Cost Center": "100392",
        "Reason for travel": "Workshop",
    },
    dir_name: str = "/Users/avosseler/Business Trips/2025/tmp",
    input_file: str = "Travel Expense Tmp.xlsx",
    output_file: str = "Travel Expense Tmp Edt.xlsx",
) -> None:
    """
    Update the travel expense Excel file with extracted invoice data.

    Args:
        result (dict): The result dictionary containing invoice entities.
        dir_name (str): Directory containing the Excel file.
        input_file (str): Name of the input Excel file.
        output_file (str): Name of the output Excel file.
        sheet_name (str): Name of the worksheet to update.
    """
    pfile = os.path.join(dir_name, input_file)
    if not os.path.exists(pfile):
        raise FileNotFoundError(f"Input file {pfile} does not exist.")
    my_logger.info(f"🤖 Updating travel expense file: {os.path.basename(pfile)}")
    wb = load_workbook(pfile, data_only=False)
    ws = wb["RKA Seite 1"]

    # --------------------------- Sheet 1 ------------------------------------
    # Overwrite the cells that feed into formulas
    ws["C2"] = trip_metadata.get("Last/First name", "Vosseler, Alexander")
    ws["E2"] = trip_metadata.get("Cost Center", "100392")  # cost center
    ws["E3"] = trip_metadata.get("Location", "Munich")  # location
    ws["C6"] = trip_metadata.get("Destination", "Barcelona")  # destination
    ws["C7"] = trip_metadata.get("Reason for travel", "Workshop")  # Reason for travel

    rowA = 19
    entries = 1
    rowB = 29
    for res in result["entities"]:
        if res["invoice_type"] == "hotel":
            ws[f"A{rowA}"] = res["checkin_date"]
            ws[f"B{rowA}"] = entries
            ws[f"C{rowA}"] = res["description"]
            ws[f"E{rowA}"] = res["total_amount"]
            rowA += 1
        else:
            ws[f"A{rowB}"] = res["issue_date"]
            ws[f"B{rowB}"] = entries
            ws[f"C{rowB}"] = res["description"]
            ws[f"E{rowB}"] = res["total_amount"]
            rowB += 1
        entries += 1

    ws[f"C{rowB+2}"] = result["rate_info"]  # Add exchange rate info

    # --------------------------- Sheet 2 ------------------------------------
    ws2 = wb["RKA Seite 2"]

    # Manually attach CalcProperties to force full recalc on load
    wb._calculation_properties = CalcProperties(fullCalcOnLoad=True)

    # Save to a new file (or overwrite)
    wb.save(os.path.join(dir_name, output_file))
    my_logger.info(
        f"✅ Updated travel expense file saved as: {os.path.join(dir_name, output_file)}"
    )
